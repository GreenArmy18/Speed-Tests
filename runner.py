import openpyxl
import speedtest
from datetime import datetime

wb = openpyxl.load_workbook("C:/Users/User/Desktop/Speed-Tests/Speed_Tests.xlsx")

sheet = wb.active

last_row = sheet.max_row

s = speedtest.Speedtest()
d = s.download() / 1000000
u = s.upload() / 1000000
servernames = []
s.get_servers(servernames)
p = s.results.ping

i = s.results.share()

sheet.cell(row=last_row + 1, column=1).value = datetime.now().strftime(
    "%d/%m/%Y %H:%M:%S"
)
sheet.cell(row=last_row + 1, column=2).value = d
sheet.cell(row=last_row + 1, column=3).value = u
sheet.cell(row=last_row + 1, column=4).value = p
sheet.cell(row=last_row + 1, column=5).value = i
wb.save("C:/Users/User/Desktop/Speed-Tests/Speed_Tests.xlsx")
